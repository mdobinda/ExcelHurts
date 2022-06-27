from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Magdalena": {
        "Settlement Date": '01/01/2017',
        "Maturity Date": '06/30/2019',
        "Rate of Interest": 0.10,
        "Price per $100 FV": 101,
        "Redemption Value": 100,
        "Payment terms": 4
    },

    "Ashok": {
        "Settlement Date": '01/01/2017',
        "Maturity Date": '06/30/2019',
        "Rate of Interest": 0.10,
        "Price per $100 FV": 101,
        "Redemption Value": 100,
        "Payment terms": 2
    }
}


wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Magdalena'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

for col in range(2, len(data['Magdalena']) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '3'})*{len(data)}"

for col in range(1, 9):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")


ws.insert_cols(8)
ws['I1'] = "Yield"


ws["I2"] = f"=YIELD({'B2'},{'C2'},{'D2'},{'E2'},{'F2'},{'G2'}, 0)"
ws["I3"] = f"=YIELD({'B3'},{'C3'},{'D3'},{'E3'},{'F3'},{'G3'}, 0)"

wb.save("NewGrades2.xlsx")
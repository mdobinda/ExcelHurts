from openpyxl import Workbook, load_workbook

wb = load_workbook('Gradez.xlsx')
ws = wb.active  # gives u active worksheet from tjhis workbook

# print(ws['A2'].value)

ws['A2'].value = "Test"
ws['A5'].value = "Kyle"
# you can also format it as this, same result
# however when you are ACCESSING you need to use .value
# saving the sheet if u make changes wb.save('Gradez.xlsx')
ws['A4'] = "Test"

# lets learn how to grab different sheets
# print(wb.sheetnames)
# print(wb['Sheet1'])

# how to create new sheet

wb.create_sheet("TestSheet")


print(wb.sheetnames)

# lets do rowssss

# ws.append(['Tim', 'Is', 'Great', '!'])
# lets do rowssss

#loop through rows and cols
# for row in range(1, 11):
#     #lets look at row 1 through 10, stop at 11.
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         # print(ws[char + str(row)].value)
#         ws[char + str(row)] = char + str(row)
#



# merging cells
ws.merge_cells("A1:D1")


#inserting rows
ws.insert_rows(7)
ws.insert_rows(10)


ws.delete_rows(10)
wb.save('tim.xlsx')
